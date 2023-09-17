from flask import Flask, render_template, jsonify
import random
import openpyxl

app = Flask(__name__)

# 엑셀 파일 생성
workbook = openpyxl.Workbook()

# 워크시트 선택
worksheet = workbook.active
worksheet.title = '자리 배치'

# 20명의 무작위 이름 생성
names = ["박경민","고아라","김리혜","김유건","김지원","박상민","백강인","석혜린","김시현","양지유","조예원","박은지","이가희",
    "이예린","이지윤","임나윤","장정현","최지휘","백정민","문요한"]


worksheet['D2'].value = '교'
worksheet['E2'].value = '탁'


# 20명의 이름을 무작위로 섞음
random.shuffle(names)


last_two_name = names[-2]
last_one_name = names[-1]
names = names[:-2]

# 자리 배치
row, col = 4, 2  # 시작 셀 위치

for name in names:
    worksheet.cell(row=row, column=col, value=name)
    col += 1
    if col > 7:  # 열이 6을 초과하면 다음 행으로 이동
        row += 2
        col = 2

worksheet.cell(row=10 , column = 4 , value = last_one_name)
worksheet.cell(row=10 , column = 5 , value = last_two_name)


workbook.save('자리배치.xlsx')


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/run-python-code', methods=['POST'])
def run_python_code():
    # 여기에 실행할 Python 코드를 추가합니다
    result = "Python 코드 실행 결과"
    return jsonify({'result': result})

if __name__ == '__main__':
    app.run(debug=True)


